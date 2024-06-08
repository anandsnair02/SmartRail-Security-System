from pathlib import Path
import tkinter as tk
from tkinter import ttk
import cv2
from PIL import Image, ImageTk
import torch
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
import csv
from datetime import datetime
import face_recognition
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import shutil
from tkinter import filedialog



# from tkinter import *
# Explicit imports to satisfy Flake8
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage



window = Tk()
window.title("SmartRail Security System")

window.geometry("1918x987")
window.configure(bg = "#101010")


canvas = tk.Canvas(
    window,
    bg="#101010",
    height=987,
    width=881,
    bd=0,
    highlightthickness=0,
    relief="ridge"
)
canvas.place(x=0, y=0)

text_variable = tk.StringVar()
text_variable.set("Dashboard")  # Initial text

canvas_text = canvas.create_text(
    194.0,
    33.0,
    anchor="nw",
    text="SmartRail Security System",  # Initial text
    fill="#FFFFFF",
    font=("Lato SemiBold", 28)
)


line_id = canvas.create_line(
    138.0,
    -1.0,
    138.0,
    986.0,
    fill="#B0B0B0"
)
canvas.tag_raise(line_id)





cap = None

def stop_running_program():
    # Release camera resources if any program is already running
    global cap
    try:
        cap.release()
        
    except AttributeError:
        pass




def change_text(new_text):
    canvas.itemconfig(canvas_text, text=new_text)


   

# Function to open dashboard
def open_dashboard():
    
 class WebcamApp:
    def __init__(self, video_labelg):
        self.video_label = video_labelg
        # Open the default camera (usually the built-in webcam)317
        self.cap = cv2.VideoCapture(0)
        video_labelg = tk.Label(window,bg="#101010")
        # Create four canvases to display frames
        self.canvas_tl = tk.Canvas(video_label, width=616, height=317.47)
        self.canvas_tl.grid(row=0, column=0)
        self.canvas_tr = tk.Canvas(video_label, width=616, height=317.47)
        self.canvas_tr.grid(row=0, column=1)
        self.canvas_bl = tk.Canvas(video_label, width=616, height=317.47)
        self.canvas_bl.grid(row=1, column=0)
        self.canvas_br = tk.Canvas(video_label, width=616, height=317.47)
        self.canvas_br.grid(row=1, column=1)
        
        # Call the show_frame function to display frames
        self.show_frame()
        
    def show_frame(self):
        # Capture frame-by-frame
        ret, frame = self.cap.read()
        if ret:
            # Convert the frame from BGR to RGB format
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            
            # Resize the frame to fit the canvas
            frame = cv2.resize(frame, (616, 317))
            
            # Convert the frame to ImageTk format
            photo = ImageTk.PhotoImage(image=Image.fromarray(frame))
            
            # Display the frame on each canvas
            self.canvas_tl.create_image(0, 0, image=photo, anchor=tk.NW)
            self.canvas_tr.create_image(0, 0, image=photo, anchor=tk.NW)
            self.canvas_bl.create_image(0, 0, image=photo, anchor=tk.NW)
            self.canvas_br.create_image(0, 0, image=photo, anchor=tk.NW)
            
            # Keep a reference to the photo to prevent it from being garbage collected
            self.canvas_tl.photo = photo
            self.canvas_tr.photo = photo
            self.canvas_bl.photo = photo
            self.canvas_br.photo = photo
        
        # Call the show_frame function after 10 milliseconds
        self.video_label.after(10, self.show_frame)
        
    def close(self):
        # Release the camera and close the window
        self.cap.release()
        self.video_label.destroy()
        global webcam_app


 app = WebcamApp(video_label)

 stop_running_program()
 change_text("Dashboard")


# Define button properties
button_text = "Dashboard"
button_font = ("Lato Regular", 16)
button_fill = "#FFFFFF"
button_position = (12.5,92.0)

# Create the button
button = tk.Button(canvas, text=button_text, font=button_font, fg=button_fill, bg="#101010", highlightthickness=0, borderwidth=0, command=open_dashboard)
button_window = canvas.create_window(button_position[0], button_position[1], anchor="nw", window=button)

# Label to display video feed
video_label = tk.Label(window,bg="#101010")
video_label.place(x=238.0, y=160.0, width=1221.0, height=593.0)
canvas.create_text(
    41.0,
    845.0,
    anchor="nw",
    text="Logout",
    fill="#B0B0B0",
    font=("Lato Regular", 16 * -1)
)



# Function to open missing person detection
def track_images():
 # Initialize face recognition
 def initialize_face_recognition():
    known_face_encodings = []
    known_face_names = []

    image_folder = "Images"
    for filename in os.listdir(image_folder):
        if filename.endswith(".jpg") or filename.endswith(".png"):
            image_path = os.path.join(image_folder, filename)
            image = face_recognition.load_image_file(image_path)
            face_encodings = face_recognition.face_encodings(image)

            if len(face_encodings) > 0:
                face_encoding = face_encodings[0]
                known_face_encodings.append(face_encoding)
                known_face_names.append(os.path.splitext(filename)[0])
            else:
                print(f"No face detected in {filename}")

    return known_face_encodings, known_face_names

 #  Function to open webcam and integrate face recognition
 def open_webcam_and_face_recognition():
    cap = cv2.VideoCapture(0)
      # Get the dimensions of the video frames
    frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

    # Function to update the video feed and perform face recognition
    def update_feed():
     
        _, framez = cap.read()
        if framez is None:
            return
        original_framez = framez.copy()
        framez = cv2.cvtColor(framez, cv2.COLOR_BGR2RGB)
        # Perform face recognition
        face_locations = face_recognition.face_locations(framez)
        face_encodings = face_recognition.face_encodings(framez, face_locations)

        for face_encoding, face_location in zip(face_encodings, face_locations):
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            if True in matches:
                first_match_index = matches.index(True)
                name = known_face_names[first_match_index]

                # Write to Excel
                write_to_excel(name)
                # Get current date and time
                current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                messagebox.showinfo("Face Detected", f"Known face detected: {name}\nDate: {current_datetime}")
                top, right, bottom, left = face_location

               # Adjust the coordinates to original frame dimensions
                top = int(top * frame_height / original_framez.shape[0])
                right = int(right * frame_width / original_framez.shape[1])
                bottom = int(bottom * frame_height / original_framez.shape[0])
                left = int(left * frame_width / original_framez.shape[1])

            top, right, bottom, left = face_location
            cv2.rectangle(framez, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(framez, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)

        # Display video feed
        
        framez = Image.fromarray(framez)
        framez = ImageTk.PhotoImage(image=framez)
        video_labelg.img = framez
        video_labelg.config(image=framez)
        video_label.after(10, update_feed)  # Update every 10 milliseconds

    update_feed()

 # Function to write data to Excel
 def write_to_excel(name):
 # Load or create the workbook
    try:
        wb = load_workbook('recognized_faces.xlsx')
    except FileNotFoundError:
        wb = Workbook()
        wb.save('recognized_faces.xlsx')
        wb = load_workbook('recognized_faces.xlsx')

    ws = wb.active

   
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")

    # Append the new data
    ws.append([name, current_date, current_time])

    wb.save('recognized_faces.xlsx')

  # Initialize known face encodings and names
 known_face_encodings, known_face_names = initialize_face_recognition()

 # Label to display video feed
 video_labelg = tk.Label(window,bg="#101010")
 video_labelg.place(x=258.0, y=190.0, width=1221.0, height=593.0)
 open_webcam_and_face_recognition()
 stop_running_program()
 change_text("Identification")



# Define button properties for missing person detection

button_text2 = "Identification"
button_font2 = ("Lato Regular", 14)
button_fill2 = "#B0B0B0"
button_position2 = (8.0,692.0)   

button2 = tk.Button(canvas, text=button_text2, font=button_font2, fg=button_fill2, bg="#101010", highlightthickness=0, borderwidth=0, command=track_images)
button_window2 = canvas.create_window(button_position2[0], button_position2[1], anchor="nw", window=button2)

def display_excel():
     # Label to display video feed
    video_labelg = tk.Label(window,bg="#101010")
    video_labelg.place(x=258.0, y=190.0, width=1221.0, height=593.0)
   
    wb = load_workbook('recognized_faces.xlsx')
    ws = wb.active

    wb.save('recognized_faces.xlsx')

    # Create a string to hold the Excel content
    excel_content = "Name\t     Date\tTime\n"

    # Populate the string with the content of the Excel sheet
    for row in ws.iter_rows(values_only=True):
        row_str = "\t".join(map(str, row)) + "\n"
        excel_content += row_str
        
    # Combine headings and Excel content
    full_content = excel_content

    video_labelg.config(text=full_content, bg="white", fg="#101010", font=("Courier", 10))

    # Update the video_label to apply changes
    video_labelg.place(x=320, y=190, width=500, height=533)
    # Update the video_label to apply changes
    video_labelg.update_idletasks()
    def stop_running_program(self):
     # Remove the specified Tkinter labels
     self.label.grid_remove()
     self.contact_label.grid_remove()
     self.age_label.grid_remove()


class ImageFileManager:
    folder_path = "Images"  
    def __init__(self, root, folder_path):
        self.root = root
        self.folder_path = folder_path
        self.image_files = self.load_images()
        self.current_index = 0

        frame = tk.Frame(root, width=600, height=533, bg="white")
        frame.place(x=850, y=190)  # Adjust the coordinates as needed

        # Place your widgets inside the frame
        self.image_label = tk.Label(frame)
        self.image_label.place(x=110, y=160)  # Adjust the coordinates as needed

        title_label = tk.Label(frame, text="LIST OF MISSING INDIVIDUALS", font=("Arial", 16, "bold"), bg="white")
        title_label.place(x=150, y=10)  #

        self.image_name_label = tk.Label(frame, text="")
        self.image_name_label.place(x=250, y=490)  # Adjust the coordinates as needed

        self.search_entry = tk.Entry(frame, width=50)
        self.search_entry.place(x=150, y=60)  # Adjust the coordinates as needed



        self.search_button = tk.Button(frame, text="Search", command=self.search_image)
        self.search_button.place(x=280, y=100)  # Adjust the coordinates as needed

        self.prev_button = tk.Button(frame, text="Previous", command=self.show_previous_image)
        self.prev_button.place(x=200, y=100)  # Adjust the coordinates as needed

        self.next_button = tk.Button(frame, text="Next", command=self.show_next_image)
        self.next_button.place(x=350, y=100)  # Adjust the coordinates as needed
    

        self.show_image()

    def load_images(self):
        image_files = [file for file in os.listdir(self.folder_path) if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif'))]
        return image_files

    def show_image(self):
        if self.image_files:
            image_path = os.path.join(self.folder_path, self.image_files[self.current_index])
            image = Image.open(image_path)
            image.thumbnail((400, 400))
            photo = ImageTk.PhotoImage(image)
            self.image_label.config(image=photo)
            self.image_label.image = photo
            self.image_name_label.config(text=self.image_files[self.current_index])

    def show_next_image(self):
        self.current_index = (self.current_index + 1) % len(self.image_files)
        self.show_image()

    def show_previous_image(self):
        self.current_index = (self.current_index - 1) % len(self.image_files)
        self.show_image()
    
    def search_image(self):
        search_term = self.search_entry.get().lower()
        if search_term:
            found_index = next((i for i, image_name in enumerate(self.image_files) if search_term in image_name.lower()), None)
            if found_index is not None:
                self.current_index = found_index
                self.show_image()
            else:
                self.image_name_label.config(text="Image not found.")

def open_reports():
    stop_roi()
    
    # Function to handle the button click event
    display_excel()
    file_manager = ImageFileManager(window,ImageFileManager.folder_path)
    stop_running_program()
    change_text("Reports")

# Create the button text
button_textq = "Reports"

# Create the button
buttonq = canvas.create_text(
    37.02,
    527.0,
    anchor="nw",
    text=button_textq,
    fill="#B0B0B0",
    font=("Lato Regular", 16)
)

# Bind the button click event to the open_reports function
canvas.tag_bind(buttonq, "<Button-1>", lambda event: open_reports())

def roi():
    model = torch.hub.load('ultralytics/yolov5', 'yolov5s', pretrained=True)
    count = 0
    cap = cv2.VideoCapture(0)
    area = [(509, 491), (206, 215), (378, 190), (1018, 265)]
    
    # Function to update the video feed
    def update_feed():
        nonlocal count
        nonlocal cap
        ret, frame = cap.read()
        if ret:
            frame = cv2.resize(frame, (1020, 500))
            results = model(frame)
            for index, row in results.pandas().xyxy[0].iterrows():
                x1 = int(row['xmin'])
                y1 = int(row['ymin'])
                x2 = int(row['xmax'])
                y2 = int(row['ymax'])
                d = (row['name'])
                cx = int(x1 + x2) // 2
                cy = int(y1 + y2) // 2
                results = cv2.pointPolygonTest(np.array(area, np.int32), ((cx, cy)), False)
                if results >= 0:
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
                    cv2.putText(frame, str(d), (x1, y1), cv2.FONT_HERSHEY_COMPLEX, 0.5, (255, 0, 0), 2)
                    cv2.circle(frame, (cx, cy), 3, (255, 0, 0), -1)
                    if d == 'person':
                        messagebox.showinfo("Person Detected", "A person has been detected!")
            cv2.polylines(frame, [np.array(area, np.int32)], True, (0, 0, 255), 2)
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(frame)
            img = ImageTk.PhotoImage(image=img)
            video_label.img = img  # Keep reference to prevent garbage collection
            video_label.config(image=img)
        window.after(10, update_feed)  # Update every 10 milliseconds
    
    update_feed()
    def stop_roi():
     cap.release()
     window.destroy()
    
def stop_roi():
    global cap
    
    video_label.destroy()

# Function to open dashboard
def tresspassing():
 roi()
 stop_running_program()

 change_text("Detection")

# Define button properties
button_text = "Detection"
button_font = ("Lato Regular", 16)
button_fill = "#B0B0B0"
button_position = (26.0, 223)   

button = tk.Button(canvas, text=button_text, font=button_font, fg=button_fill, bg="#101010", highlightthickness=0, borderwidth=0, command=tresspassing)
button_window = canvas.create_window(button_position[0], button_position[1], anchor="nw", window=button)

class TakeImages(tk.Frame):
    def rename_image(self, source_file, new_filename, destination_folder):
        _, extension = os.path.splitext(source_file)
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)
        destination_file_path = os.path.join(destination_folder, f"{new_filename}{extension}")
        shutil.copyfile(source_file, destination_file_path)
        print(f"Renamed {source_file} to {new_filename}{extension} and moved to {destination_folder}")

    def select_and_rename(self):
        file_path = filedialog.askopenfilename(title="Select Image", filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.gif")])
        if file_path:
            new_filename = self.name_entry.get()
            self.rename_image(file_path, new_filename, self.save_folder)
    
    def __init__(self, master):
        super().__init__(master)

        self.save_folder = "Images"

          # Label and Entry for image name
    
          # Label and Entry for image name
        self.label = tk.Label(self.master, text="Enter the Name:")
        self.label.place(x=63, y=10)

        self.name_entry = tk.Entry(self.master, width=20)
        self.name_entry.place(x=160, y=10)

        # Capture Image button
        self.capture_button = tk.Button(self.master, text="Capture Image", command=self.capture_image)
        self.capture_button.place(x=300, y=10)

        # Rename Image button
        self.rename_button = tk.Button(self.master, text="Browse Image", command=self.select_and_rename)
        self.rename_button.place(x=300, y=40)

        # Contact Info Label and Entry
        self.contact_label = tk.Label(self.master, text="Contact Info:")
        self.contact_label.place(x=78, y=40)

        self.contact_entry = tk.Entry(self.master, width=20)
        self.contact_entry.place(x=160, y=40)

        # Age Label and Entry
        self.age_label = tk.Label(self.master, text="Age:")
        self.age_label.place(x=123, y=70)

        self.age_entry = tk.Entry(self.master, width=20)
        self.age_entry.place(x=160, y=70)

        # Submit Button
        self.submit_button = tk.Button(self.master, text="Submit", command=self.capture_info_and_stop)
        self.submit_button.place(x=180, y=100)
        
        self.tree = ttk.Treeview(master, columns=("Name", "Contact Info", "Age", "Timestamp"), show="headings")
        self.tree.heading("Name", text="Name")
        self.tree.heading("Contact Info", text="Contact Info")
        self.tree.heading("Age", text="Age")
        self.tree.heading("Timestamp", text="Timestamp")
        self.tree.place(in_=self, x=0, y=0, width=530,height=500)



    def load_excel_data(self):
        filename = "form_data.xlsx"
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                self.tree.insert("", "end", values=row)
            wb.close()
    def capture_info_and_stop(self):
        self.capture_info()  # Call capture_info first
        #self.stop_running_programa()
    def stop_running_programa(self):
         self.label.grid_remove()
         self.name_entry.grid_remove()
         self.capture_button.grid_remove()
         self.rename_button.grid_remove()
         self.contact_label.grid_remove()
         self.contact_entry.grid_remove()
         self.age_label.grid_remove()
         self.age_entry.grid_remove()
         self.submit_button.grid_remove()
         self.tree.place_forget()
         self.grid_remove()
     

    def capture_info(self):
        name = self.name_entry.get()
        if name == "":
            messagebox.showerror("Error", "Please enter a name for the image.")
            return
        
        contact_info = self.contact_entry.get()
        age = self.age_entry.get()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Saving data to Excel
        filename = f"{self.save_folder}/form_data.xlsx"
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Form Data"
            ws.append(["Name", "Contact Info", "Age", "Timestamp"])
        else:
            wb = load_workbook(filename)
            ws = wb.active
        ws.append([name, contact_info, age, timestamp])
        wb.save(filename)
        print("Data saved successfully.")
        
        # Reload Excel data
        self.tree.delete(*self.tree.get_children())  # Clear existing data
        self.load_excel_data()
   

    def capture_image(self):
        name = self.name_entry.get()
        if name == "":
            messagebox.showerror("Error", "Please enter a name for the image.")
            return

        # Create save folder if it doesn't exist
        if not os.path.exists(self.save_folder):
            os.makedirs(self.save_folder)

        # Initialize webcam
        cap = cv2.VideoCapture(0)
        cap.set(3, 640)  # Set width
        cap.set(4, 480)  # Set height

        ret, frame = cap.read()

        # Save the image
        image_path = os.path.join(self.save_folder, f'{name}.jpg')
        cv2.imwrite(image_path, frame)
        print(f"Image saved as {name}.jpg")

        # Release the webcam and close OpenCV windows
        cap.release()
        cv2.destroyAllWindows()
        messagebox.showinfo("Success", f"Image saved as {name}.jpg")
       


def Activities():
    stop_roi()
     # Label to display video feed
    video_labelg = tk.Label(window,bg="#101010")
    video_labelg.place(x=258.0, y=190.0, width=1221.0, height=593.0)
  
    # Label to display video feed
    
    # Create an instance of TakeImages inside the video_label
    take_images_instance = TakeImages(video_labelg)
    take_images_instance.pack()
    stop_running_program()
    change_text("Activities")


# Define button properties
button_text = "Activities"
button_font = ("Lato Regular", 16)
button_fill = "#B0B0B0"
button_position = (24.0, 380.0)

# Create the button
button = tk.Button(canvas, text=button_text, font=button_font, fg=button_fill, bg="#101010", highlightthickness=0, borderwidth=0,  command=Activities)
button_window = canvas.create_window(button_position[0], button_position[1], anchor="nw", window=button)





#window.resizable(False, False)
window.mainloop()
